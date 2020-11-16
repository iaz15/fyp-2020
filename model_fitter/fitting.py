"""
Fit Multiple Data Sets
======================

Fitting multiple (simulated) Gaussian data sets simultaneously.

All minimizers require the residual array to be one-dimensional. Therefore, in
the ``objective`` we need to ```flatten``` the array before returning it.

Adapted from: https://lmfit.github.io/lmfit-py/examples/example_fit_multi_datasets.html
Check out https://newville.github.io/asteval/ - asteval an alternative to eval

Extra resources, refer to:
    https://github.com/lmfit/lmfit-py/blob/master/lmfit/models.py
    https://github.com/lmfit/lmfit-py/blob/master/lmfit/model.py
    https://lmfit.github.io/lmfit-py/model.html
"""
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from numba import jit
import time
import json
import os
import copy

from lmfit import Parameters, report_fit
import lmfit
import collections
import itertools
import openpyxl
import datetime

from models import (StaticExperimentalConditions, FrictionModelParametersLiquid,
    FrictionModelLiquid, FrictionLiquidVariableConditionsData)
from models import (FrictionModelParametersLiquidSolid, Lubricant,
    FrictionModelLiquidSolid, FrictionLiquidSolidVariableConditionsData)
from sample_models import (GaussModel, GaussParameters, GaussStaticConditions, GaussVariableConditionsData)

from matplotlib.widgets import Slider, Button, CheckButtons

class Fitter():

    def __init__(self, model) -> None:
        self.model = model
        self.datasets = []
        self.param_vary_names = Fitter.get_all_properties(model.parameters)
        self.param_static_names = []
        self.param_shared_names = []
        self.variable_conditions_names = []
        self.fit_params = None
        self.plot_list = []
        self.recent_plot = None

        if isinstance(model, GaussModel):
            pass
        elif isinstance(model, FrictionModelLiquid):
            pass
        elif isinstance(model, FrictionModelLiquidSolid):
            pass
        else:
            raise Exception("You need to pass in a valid Model object! (Pick from Gauss or Friction Models)")

    def print_all_params(self):
        print("")
        print(f"Variable parameters: {self.param_vary_names}")
        print(f"Static parameters: {self.param_static_names}")
        print(f"Shared parameters: {self.param_shared_names}")
        print(f"Variable conditions: {self.variable_conditions_names}")

    def make_param_vary(self, param_name, verbose=True):
        if param_name in Fitter.get_all_properties(self.model.parameters):

            if param_name in self.param_vary_names:
                if verbose == True:
                    print(f"Parameter '{param_name}' is already in vary parameters")
                if param_name in self.param_shared_names:
                    print(f"Parameter '{param_name}' is also in shared parameters, proceeding to unshare")
                    for i in range(1, len(self.datasets)+1):
                        self.fit_params[f'{param_name}_{i}'].vary = True
                        self.fit_params[f'{param_name}_{i}'].expr = None
                    self.param_shared_names.remove(param_name)
            else:
                if verbose == True:
                    print(f"Parameter '{param_name}' is in the static parameters, proceeding to make variable")
                self.param_static_names.remove(param_name)
                self.param_vary_names.append(param_name)

                for i in range(1, len(self.datasets)+1):
                    self.fit_params[f'{param_name}_{i}'].vary = True
                    self.fit_params[f'{param_name}_{i}'].expr = None
        else:
            error_message = f"Parameter {param_name} is not a valid parameter name"
            raise Exception(error_message)

    def make_param_static(self, param_name, verbose=True):
        if param_name in Fitter.get_all_properties(self.model.parameters):

            if param_name in self.param_static_names:
                if verbose == True:
                    print(f"    Parameter '{param_name}' is already in static parameters")
            else:
                if verbose == True:
                    print(f"    Parameter '{param_name}' is in the vary parameters, proceeding to make static")
                self.param_vary_names.remove(param_name)
                self.param_static_names.append(param_name)

                for i in range(1, len(self.datasets)+1):
                    self.fit_params[f'{param_name}_{i}'].vary = False
                    self.fit_params[f'{param_name}_{i}'].expr = None

                if param_name in self.param_shared_names:
                    if verbose == True:
                        print(f"    Parameter '{param_name}' is also in the shared parameters, removing from shared list")
                    self.param_shared_names.remove(param_name)
        else:
            error_message = f"Parameter {param_name} is not a valid parameter name"
            raise Exception(error_message)

    def make_param_shared(self, param_name, verbose=True):
        if param_name in Fitter.get_all_properties(self.model.parameters):

            if param_name in self.param_shared_names:
                if verbose == True:
                    print(f"    Parameter '{param_name}' is already in shared parameters")
            elif param_name in self.param_vary_names:
                if verbose == True:
                    print(f"    Parameter '{param_name}' is in the vary parameters, proceeding to make shared")
                self.param_shared_names.append(param_name)

                for i in range(1, len(self.datasets)+1):
                    if i == 1:
                        self.fit_params[f'{param_name}_{i}'].vary = True
                        continue
                    self.fit_params[f'{param_name}_{i}'].expr = f"{param_name}_1"
            else:
                if verbose == True:
                    print(f"    Parameter '{param_name}' is in the static parameters, proceeding to make vary & shared")
                self.param_static_names.remove(param_name)
                self.param_vary_names.append(param_name)
                self.param_shared_names.append(param_name)
        else:
            error_message = f"Parameter {param_name} is not a valid parameter name"
            raise Exception(error_message)

    def make_param_unshared(self, param_name, verbose=True):
        if param_name in Fitter.get_all_properties(self.model.parameters):

            if param_name in self.param_shared_names:
                if verbose == True:
                    print(f"    Parameter '{param_name}' is in the shared parameters, proceeding to remove from shared")
                self.param_shared_names.remove(param_name)
            else:
                if verbose == True:
                    print(f"    Parameter '{param_name}' is not in the shared parameters")
        else:
            error_message = f"Parameter {param_name} is not a valid parameter name"
            raise Exception(error_message)

    def add_dataset(self, variable_conditions_data, verbose=True):

        if isinstance(variable_conditions_data, GaussVariableConditionsData):
            self.datasets.append(variable_conditions_data)

            if not self.variable_conditions_names:
                self.variable_conditions_names = Fitter.get_all_properties_ordered(self.datasets[0])
                self.variable_conditions_names.remove('x_data')
                self.variable_conditions_names.remove('y_data')

            if verbose == True:
                print(f"GaussVariableConditionsData passed in. Datasets = {len(self.datasets)}")
        elif isinstance(variable_conditions_data, FrictionLiquidVariableConditionsData):
            self.datasets.append(variable_conditions_data)

            if not self.variable_conditions_names:
                self.variable_conditions_names = Fitter.get_all_properties_ordered(self.datasets[0])
                self.variable_conditions_names.remove('x_data')
                self.variable_conditions_names.remove('y_data')

            if verbose == True:
                print(f"FrictionLiquidVariableConditionsData passed in. Datasets = {len(self.datasets)}")
        elif isinstance(variable_conditions_data, FrictionLiquidSolidVariableConditionsData):
            self.datasets.append(variable_conditions_data)

            if not self.variable_conditions_names:
                self.variable_conditions_names = Fitter.get_all_properties_ordered(self.datasets[0])
                self.variable_conditions_names.remove('x_data')
                self.variable_conditions_names.remove('y_data')

            if verbose == True:
                print(f"FrictionLiquidSolidVariableConditionsData passed in. Datasets = {len(self.datasets)}")
        else:
            raise Exception("A VariableConditionsData Class needs to be passed in")

    def manual_fitting(self, prediction_range):

        # https://towardsdatascience.com/intro-to-dynamic-visualization-with-python-animations-and-interactive-plots-f72a7fb69245

        self.manual_fig, self.manual_ax = plt.subplots(figsize=(15, 7.37))

        self.manual_ax.set_title(f"{self.model.model_name} predictions", fontsize=12)
        self.manual_ax.set_ylabel("Coefficient of Friction")
        self.manual_ax.set_xlabel("Sliding Distance / mm")
        self.manual_fig.subplots_adjust(top=0.95, bottom=0.45, right=0.48, left=0.05)

        # We have the datasets stored in self.datasets

        # For each dataset, we want to plot
        self.sliders = []
        self.plot_list = []
        self.variable_conditions_names_values = []

        colours = itertools.cycle(["tab:blue", "tab:orange", "tab:green", "tab:red", "tab:purple", "tab:brown", "tab:pink", "tab:gray", "tab:olive", "tab:cyan"])
        for dataset in self.datasets:

            labels_list = []
            ordered_dict = collections.OrderedDict()
            for _, variable_condition in enumerate(self.variable_conditions_names, 1):
                labels_list.append(f'{variable_condition[:2]} = {round(getattr(dataset, variable_condition), 4)}')
                ordered_dict[variable_condition] = getattr(dataset, variable_condition)

            textstr = ' '.join(labels_list)

            # TODO: Remember the model prediction function should be general, therefore it should just be 'predict'
            current_colour = next(colours)
            l, = plt.plot(prediction_range, self.model.predict(prediction_range,
                **ordered_dict), label=textstr, color=current_colour)
            plt.plot(dataset.x_data, dataset.y_data, linestyle="--", color=current_colour)

            self.variable_conditions_names_values.append(ordered_dict)
            self.plot_list.append(l)

        for idx, param_name in enumerate(self.param_shared_names):
            self.add_slider(idx, param_name, prediction_range, self.fit_params[f'{param_name}_1'].value,
                self.fit_params[f'{param_name}_1'].min, self.fit_params[f'{param_name}_1'].max, )

        self.manual_ax.legend(loc='lower left', bbox_to_anchor=(0, -0.7), ncol=1, framealpha = 0.2)

        # Make checkbuttons with all plotted lines with correct visibility
        self.rax = plt.axes([0.37, 0.14, 0.11, 0.2])
        self.labels = [f'Plot {idx}' for idx, line in enumerate(self.plot_list, 1)]
        self.visibility = [line.get_visible() for line in self.plot_list]
        self.check = CheckButtons(self.rax, self.labels, self.visibility)

        self.check.on_clicked(self.toggle_visibility)

        for param_vary in self.param_vary_names:
            self.fit_params[f'{param_vary}_1'].min
            self.fit_params[f'{param_vary}_1'].max

        # Add a button for resetting the parameters
        reset_button_ax = self.manual_fig.add_axes([0.86, 0.0271, 0.1, 0.04])
        reset_button = Button(reset_button_ax, 'Reset', hovercolor='0.975')

        save_params_ax = self.manual_fig.add_axes([0.70, 0.0271, 0.15, 0.04])
        save_button = Button(save_params_ax, 'Save Parameters', hovercolor='0.975')

        reset_button.on_clicked(self.reset_button_on_clicked)
        save_button.on_clicked(self.save_button_on_clicked)

        plt.show()

    def reset_button_on_clicked(self, mouse_event):
        for slider in self.sliders:
            slider.reset()

    def save_button_on_clicked(self, mouse_event):
        print("Save button clicked! Does nothing right now.")

    def add_slider(self, idx, param_name, prediction_range, valinit, min, max):
        # https://stackoverflow.com/questions/32737427/matplotlib-add-sliders-to-a-figure-dynamically
        spacing = 0.05
        ax = self.manual_fig.add_axes([0.55, 0.92-spacing*idx, 0.4, 0.03])
        slider = Slider(ax, param_name, min, max, valinit=valinit)
        self.sliders.append(slider)

        def update(val):
            # print("=== Called ===")
            setattr(self.model, param_name, val)
            for idx, _ in enumerate(self.datasets):
                setattr(self.model.parameters, param_name, val)
                # print(param_name, val)
                # print(self.variable_conditions_names_values[idx])
                self.plot_list[idx].set_ydata(self.model.predict(prediction_range, **self.variable_conditions_names_values[idx]))
                self.manual_fig.canvas.draw_idle()

        self.sliders[idx].on_changed(update)

    def toggle_visibility(self, label):
        index = self.labels.index(label)
        self.plot_list[index].set_visible(not self.plot_list[index].get_visible())
        plt.draw()

    def automatic_fit(self, minimizer_kwargs=None, save_results=False, show_plots=False):
        print("Starting automatic fitting")

        if minimizer_kwargs == None:
            minimizer_kwargs = {'method':'leastsq', 'maxfev': 10000}

        # Extract out x and data (np.ndarry of np.ndarrays)
        data = []
        x_array = []
        for dataset in self.datasets:
            data.append(dataset.y_data)
            x_array.append(dataset.x_data)

        data = np.array(data)
        x_array = np.array(x_array)

        # fit_params = self.fit_params

        # Try: Truncated newton, dual annealing, differential evolution. There are many more available
        z0 = time.perf_counter()
        mini = lmfit.Minimizer(self.objective, self.fit_params, fcn_args=(x_array, data), fcn_kws={})
        result = mini.minimize(**minimizer_kwargs)

        # Prints the final values & the time elapsed in seconds
        print(f"Function Evaluations: {result.nfev} evaluations")
        print(f"Time Elapsed: {time.perf_counter() - z0} seconds")

        if save_results:
            self._save_results(result)

        if show_plots == True:
            self._plot_results(result)
            plt.show()

        return result

    def _save_results(self, result):
        # Parameters suffixed for accessing fit parameters objects
        param_suffixed = [param_shared + "_1" for param_shared in self.param_shared_names]

        result_dict = collections.OrderedDict()
        result_dict['method'] = result.method
        result_dict['num_datapoints'] = result.ndata
        result_dict['function_evaluations'] = result.nfev
        result_dict['datetime'] = datetime.datetime.now().strftime("%d/%m/%Y, %H:%M:%S")
        result_dict['statistics'] = collections.OrderedDict()
        result_dict['parameters'] = collections.OrderedDict()

        for param in param_suffixed:
            sum_residuals_squared = np.sum(result.residual ** 2)

            result_dict['parameters'][param] = {}
            result_dict['parameters'][param]['initial_value'] = result.params[param].init_value
            result_dict['parameters'][param]['final_value'] = result.params[param].value
            result_dict['parameters'][param]['min'] = result.params[param].min
            result_dict['parameters'][param]['max'] = result.params[param].max

            result_dict['statistics']['aic'] = result.aic
            result_dict['statistics']['bic'] = result.bic
            result_dict['statistics']['chisqr'] = result.chisqr
            result_dict['statistics']['sum_residuals_squared'] = sum_residuals_squared

        json_num = 0
        while os.path.exists(f"{self.model.model_name}_results_{json_num}.json"):
            json_num += 1

        fig, _ = self._plot_results(result)
        fig.savefig(f'{self.model.model_name}_results_{json_num}.png')
        plt.close(fig)

        with open(f'{self.model.model_name}_results_{json_num}.json', 'w') as file:
            json.dump(result_dict, file, indent=4)

        return None

    def _plot_results(self, result):
        ### Plot the results
        fig, ax = plt.subplots()

        temp_list = []
        for i in range(len(self.datasets)):
            temp_list.append(i)

        colours = itertools.cycle(["tab:blue", "tab:orange", "tab:green", "tab:red", "tab:purple", "tab:brown", "tab:pink", "tab:gray", "tab:olive", "tab:cyan"])

        for idx, dataset in enumerate(self.datasets):
            current_colour = next(colours)
            y_fit = self.compute_y(result.params, idx, dataset.x_data)
            ax.plot(dataset.x_data, dataset.y_data, '--', color=current_colour)
            ax.plot(dataset.x_data, y_fit, '-', color=current_colour)

        return fig, ax

    def create_params(self, verbose=False, **kwargs):

        if not self.datasets:
            print("Aborting creating parameters, add a dataset first")
            return None

        print(f'There are {len(self.datasets)} datasets')

        fit_params = Parameters()

        for idx, dataset in enumerate(self.datasets, 1):

            if verbose==True:
                print("")
                print(f"== Dataset: {idx} ==")
            for variable_condition in self.variable_conditions_names:
                if verbose==True:
                    print(variable_condition, getattr(dataset, variable_condition))

                fit_params.add(f'{variable_condition}_{idx}', value=getattr(dataset, variable_condition) , vary=False)

            # Lets say we want amp and cen to change, but sig to stay constant
            # We know the initial values of amp, cen and sig from the model.parameters

            # First add amp, cen and sig and enforce values (self.param_vary_names, self.param_static_names)
            for param in self.param_vary_names:
                if verbose == True:
                    print("Variable Parameter:", param, getattr(self.model.parameters, param))
                fit_params.add(f'{param}_{idx}', value=getattr(self.model.parameters, param))

            for param in self.param_static_names:
                if verbose == True:
                    print("Static Parameter:", param, getattr(self.model.parameters, param))
                fit_params.add(f'{param}_{idx}', value=getattr(self.model.parameters, param), vary=False)

            for param in self.param_shared_names:
                if verbose == True:
                    print("Shared Parameter:", param, getattr(self.model.parameters, param))
                if idx == 1:
                    pass
                else:
                    fit_params.add(f'{param}_{idx}', value=getattr(self.model.parameters, param))
                    fit_params[f"{param}_{idx}"].expr = f"{param}_1"

        self.fit_params = fit_params

    def set_param_range(self, name, min, max) -> None:
        if name in self.param_vary_names:

            for idx in range(1, len(self.datasets) + 1):
                self.fit_params[f'{name}_{idx}'].set(min=min, max=max)

        elif name in self.param_static_names:
            print(f"Parameter: {name} not in variable parameters but in static parameters. Cannot set range.")
        else:
            print(f"Parameter: {name} not in variable parameters or static parameters. Cannot set range.")

    @staticmethod
    def get_all_properties(object_) -> list:
        parameters = [a for a in dir(object_) if not a.startswith('__') and not \
            callable(getattr(object_, a))]

        return parameters

    @staticmethod
    def get_all_properties_ordered(object_) -> list:
        return list(object_.__dict__.keys())

    def objective(self, params, x, data) -> np.ndarray:
        """Calculate total residual for fits of Gaussians to several data sets."""
        ndata, *_ = data.shape
        resid = 0.0*data[:]

        # make residual per data set
        for i in range(ndata):
            resid[i] = data[i] - self.compute_y(params, i, x[i]).flatten()

        # Now flatten this to a 1D array, as minimize() requires a 1D array
        # to be optimised. Note that .flatten() doesn't work here because
        # the arrays are of different lengths.
        flattened_output = np.hstack(resid)

        return flattened_output

    def compute_y(self, params, i, x):
        """Calculate Gaussian lineshape from parameters for data set."""

        for variable_parameter in self.param_vary_names:
            setattr(self.model.parameters, variable_parameter, params[f'{variable_parameter}_{i+1}'].value)

        # This dictionary approach works if it not compiled into machine code. It is the most flexible method.
        # variable_conditions_values = {}
        # for variable_condition in self.variable_conditions_names:
        #     variable_conditions_values[variable_condition] = params[f'{variable_condition}_{i+1}'].value
        # result = self.model.predict(x, **variable_conditions_values)

        # This works only for solid + liquid friction model. This works with machine compiled python code.
        # Requires the names in the list to be the same order as the way the inputs are taken in
        # for the predict function.
        variable_conditions_values = []
        for variable_condition in self.variable_conditions_names:
            variable_conditions_values.append(params[f'{variable_condition}_{i+1}'].value)

        result = self.model.predict(x, *variable_conditions_values)

        return result

def fit_gaussian():
    parameters = GaussParameters(100, 0.4, 0.3)
    static_conditions = GaussStaticConditions('P20', 0.3, 'AA7075', 0.7)
    model = GaussModel(parameters, static_conditions)

    temperature = 450; force =  8; pressure = 0.34; speed = 50; lubricant_thickness = 25

    # model.plot_results(plotting_range, temperature, force, pressure, speed, lubricant_thickness)

    ### SAMPLE EXTACTION FROM DB ###
    variable_conditions = [
        {'temperature': 450, 'force': 8, 'pressure': 0.34, 'speed': 74, 'lubricant_thickness': 25},
        {'temperature': 300, 'force': 8, 'pressure': 0.34, 'speed': 76, 'lubricant_thickness': 25},
        {'temperature': 200, 'force': 8, 'pressure': 0.34, 'speed': 80, 'lubricant_thickness': 25},
        {'temperature': 100, 'force': 8, 'pressure': 0.34, 'speed': 82, 'lubricant_thickness': 25},
        {'temperature': 80, 'force': 8, 'pressure': 0.34, 'speed': 86, 'lubricant_thickness': 25}
    ]

    parameters_extracted_db = [{'amp': 5, 'cen': -0.20, 'sig': 0.25}]

    ### SIMULATE DATASETS ###
    # Create five simulated Gaussian data sets
    x = np.linspace(-1, 2, 151)
    x_array = []
    data = []
    for i in np.arange(5):

        amp = parameters_extracted_db[0]['amp'] + 9.50*np.random.rand()
        cen = parameters_extracted_db[0]['cen'] + 1.20*np.random.rand()
        sig = parameters_extracted_db[0]['sig'] + 0.03*np.random.rand()

        temperature = variable_conditions[i]['temperature']
        force = variable_conditions[i]['force']
        pressure = variable_conditions[i]['pressure']
        speed = variable_conditions[i]['speed']
        lubricant_thickness = variable_conditions[i]['lubricant_thickness']

        parameters = GaussParameters(amp, cen, sig)
        conditions = GaussStaticConditions('P20', 0.3, 'AA7075', 0.7)
        model = GaussModel(parameters, conditions)

        dat = model.predict(x, temperature, force, pressure, speed,
            lubricant_thickness) + \
            np.random.normal(size=x.size, scale=0.1)

        x_array.append(x)
        data.append(dat)

    ###############################################################################
    ### START OF PROCESS FOR USER ###

    dataset1 = GaussVariableConditionsData(variable_conditions[0]['temperature'],
        variable_conditions[0]['force'], variable_conditions[0]['pressure'],
        variable_conditions[0]['speed'],  variable_conditions[0]['lubricant_thickness'],
        x_array[0], data[0])

    dataset2 = GaussVariableConditionsData(variable_conditions[1]['temperature'],
        variable_conditions[1]['force'], variable_conditions[1]['pressure'],
        variable_conditions[1]['speed'],  variable_conditions[1]['lubricant_thickness'],
        x_array[1], data[1])

    dataset3 = GaussVariableConditionsData(variable_conditions[2]['temperature'],
        variable_conditions[2]['force'], variable_conditions[2]['pressure'],
        variable_conditions[2]['speed'],  variable_conditions[2]['lubricant_thickness'],
        x_array[2], data[2])

    dataset4 = GaussVariableConditionsData(variable_conditions[3]['temperature'],
        variable_conditions[3]['force'], variable_conditions[3]['pressure'],
        variable_conditions[3]['speed'],  variable_conditions[3]['lubricant_thickness'],
        x_array[3], data[3])

    dataset5 = GaussVariableConditionsData(variable_conditions[4]['temperature'],
        variable_conditions[4]['force'], variable_conditions[4]['pressure'],
        variable_conditions[4]['speed'],  variable_conditions[4]['lubricant_thickness'],
        x_array[4], data[4])

    # Create the gauss model
    parameters = GaussParameters(parameters_extracted_db[0]['amp'], parameters_extracted_db[0]['cen'], parameters_extracted_db[0]['sig'])
    static_conditions = GaussStaticConditions('P20', 0.3, 'AA7075', 0.7)
    gauss_model = GaussModel(parameters, static_conditions)

    # Create the fitting model for gauss model
    gauss_model_fitter_1 = Fitter(gauss_model)
    gauss_model_fitter_1.add_dataset(dataset1, verbose=False)
    gauss_model_fitter_1.add_dataset(dataset2, verbose=False)
    gauss_model_fitter_1.add_dataset(dataset3, verbose=False)
    gauss_model_fitter_1.add_dataset(dataset4, verbose=False)
    gauss_model_fitter_1.add_dataset(dataset5, verbose=False)

    ### Initialise parameters
    gauss_model_fitter_1.create_params(verbose=False)

    gauss_model_fitter_1.print_all_params()

    gauss_model_fitter_1.set_param_range('amp', min=0.0, max=100.0)
    gauss_model_fitter_1.set_param_range('cen', min=-2.0, max=2.0)
    gauss_model_fitter_1.set_param_range('sig', min=0.01, max=3.0)

    gauss_model_fitter_1.make_param_shared("sig")
    gauss_model_fitter_1.make_param_shared("amp")
    gauss_model_fitter_1.make_param_vary("amp", verbose=True)

    gauss_model_fitter_1.fit_params.pretty_print()

    out = gauss_model_fitter_1.automatic_fit()
    report_fit(out.params)

    plt.figure()
    for i in range(len(gauss_model_fitter_1.datasets)):
        y_fit = gauss_model_fitter_1.compute_y(out.params, i, x_array[i])
        plt.plot(x_array[i], data[i], 'o', x_array[i], y_fit, '-')
    plt.show()

def fit_liquid():
    ### LIQUID ONLY MODEL ###
    print('== Friction Model Liquid Only ==')

    lubricant_1 = Lubricant(1, 'ZEPF', True, False, 1.125)

    ### Extracted from database
    condition_1 = StaticExperimentalConditions(lubricant=lubricant_1, pin_material='P20',
        pin_roughness=0.3, blank_material='AA7075', blank_roughness=0.7, blank_thickness=2,
        coating_material='None', coating_thickness=0, coating_roughness=0)

    # k2 should be ~2.89
    parameters_1 = FrictionModelParametersLiquid(mu0_lubricated = 1.69073,
        Q_lubricated = 9141.50683, mu0_dry = 10.94225, Q_dry = 9368.85126,
        eta_0 = 0.12, Q_eta = 11930, lambda_1 = 25.0, lambda_2 = 0.8, c = 0.1,
        k_1 = 1.2, k_2 = 3, k_3 = 4.58)

    ### Create the model
    friction_model_liquid = FrictionModelLiquid(condition_1, parameters_1)

    ### Extracted from database
    variable_conditions = [
        {'temperature': 250, 'force': 5, 'pressure': 0.34, 'speed': 50, 'lubricant_thickness': 25},
        {'temperature': 250, 'force': 5, 'pressure': 0.34, 'speed': 100, 'lubricant_thickness': 25}
    ]

    data = []
    x = []

    cof1 = np.array([0.1832, 0.21344, 0.214, 0.16829, 0.20092, 0.21337, 0.2154, 0.20336, 0.21587, 0.23699, 0.23952, 0.21664, 0.2014, 0.19174, 0.18813, 0.18991, 0.18264, 0.17702, 0.1875, 0.20547, 0.20907, 0.19216, 0.18327, 0.1949, 0.19512, 0.21268, 0.27634, 0.34194, 0.42582, 0.56496, 0.70509, 0.89468, 1.03188])
    cof2 = np.array([0.21865, 0.19783, 0.20394, 0.20249, 0.21899, 0.21676, 0.28037, 0.37011, 0.52885, 0.72599, 0.91727, 1.0663])
    sd1 = np.array([0, 1.875, 3.75, 5.625, 7.5, 9.375, 11.25, 13.125, 15, 16.875, 18.75, 20.625, 22.5, 24.375, 26.25, 28.125, 30, 31.875, 33.75, 35.625, 37.5, 39.375, 41.25, 43.125, 45, 46.875, 48.75, 50.625, 52.5, 54.375, 56.25, 58.125, 60])
    sd2 = np.array([0, 1.875, 3.75, 5.625, 7.5, 9.375, 11.25, 13.125, 15, 16.875, 18.75, 20.625])

    x.append(sd1)
    x.append(sd2)
    data.append(cof1)
    data.append(cof2)

    dataset1 = FrictionLiquidVariableConditionsData(variable_conditions[0]['temperature'],
        variable_conditions[0]['force'], variable_conditions[0]['pressure'],
        variable_conditions[0]['speed'],  variable_conditions[0]['lubricant_thickness'],
        x[0], data[0])

    dataset2 = FrictionLiquidVariableConditionsData(variable_conditions[1]['temperature'],
        variable_conditions[1]['force'], variable_conditions[1]['pressure'],
        variable_conditions[1]['speed'],  variable_conditions[1]['lubricant_thickness'],
        x[1], data[1])

    friction_liquid_model_fitter_1 = Fitter(friction_model_liquid)

    ### Add datasets
    friction_liquid_model_fitter_1.add_dataset(dataset1, verbose=True)
    friction_liquid_model_fitter_1.add_dataset(dataset2, verbose=True)

    ### Create parameters
    friction_liquid_model_fitter_1.create_params(verbose=True)

    friction_liquid_model_fitter_1.make_param_static('Q_dry')
    friction_liquid_model_fitter_1.make_param_static('Q_lubricated')
    friction_liquid_model_fitter_1.make_param_static('Q_eta')

    friction_liquid_model_fitter_1.make_param_static('mu0_dry')
    friction_liquid_model_fitter_1.make_param_static('mu0_lubricated')
    friction_liquid_model_fitter_1.make_param_static('eta_0')

    ### Set ranges
    friction_liquid_model_fitter_1.set_param_range('k_1', min=1, max=3)
    friction_liquid_model_fitter_1.set_param_range('k_2', min=2, max=3.5)
    friction_liquid_model_fitter_1.set_param_range('k_3', min=1, max=10.0)
    friction_liquid_model_fitter_1.set_param_range('lambda_1', min=0, max=10000)
    friction_liquid_model_fitter_1.set_param_range('lambda_2', min=0.2, max=3.0)
    friction_liquid_model_fitter_1.set_param_range('c', min=0.001, max=0.01)

    ### Make params shared
    friction_liquid_model_fitter_1.make_param_shared("k_1")
    friction_liquid_model_fitter_1.make_param_shared("k_2")
    friction_liquid_model_fitter_1.make_param_shared("k_3")
    friction_liquid_model_fitter_1.make_param_shared("lambda_1")
    friction_liquid_model_fitter_1.make_param_shared("lambda_2")
    friction_liquid_model_fitter_1.make_param_shared("c")

    friction_liquid_model_fitter_1.fit_params.add('blank_roughness', value=0.7, vary=False)
    friction_liquid_model_fitter_1.fit_params.add('delta', value=10, min=8, max=16)
    friction_liquid_model_fitter_1.fit_params['lambda_1_1'].expr = 'delta ** (1.0 / lambda_2_1) / blank_roughness'
    print(friction_liquid_model_fitter_1.fit_params.pretty_print())

    ### Manual fitting
    friction_liquid_model_fitter_1.manual_fitting(x[0])

    ### Automatic fitting
    out = friction_liquid_model_fitter_1.automatic_fit()
    report_fit(out.params)

    ### Plot the results
    plt.figure()
    for i in range(len(friction_liquid_model_fitter_1.datasets)):
        y_fit = friction_liquid_model_fitter_1.compute_y(out.params, i, x[i])
        plt.plot(x[i], data[i], 'o', x[i], y_fit, '-')
    plt.show()

def fit_liquid_solid():
    ### NOTE: CHOOSE THE EXPERIMENTS YOU WANT FROM HERE
    # Experiment ID 1 will be invalid as 220 is not one that doesn't has cof specified
    # Specify the workbook_name containing the friction data

    ### User Input. You should know the number of experiments in the workbook beforehand
    ### Indexing starts at 1 for these experiments
    optimisation_results_folder = "optimisation_results"
    workbook_name = "liquid_solid_model_tests.xlsx"
    chosen_ids = [2]
    chosen_ids = [2,3,4,5,6,7,8]

    lubricant_2 = Lubricant(id=2, name='Omega 35', includes_liquid=True, includes_solid=True,
        density_liquid=1.125, avg_hardness_solid=0.088)

    condition_2 = StaticExperimentalConditions(lubricant=lubricant_2, pin_material='P20',
        pin_roughness=0.3, blank_material='AA7075', blank_roughness=0.7,
        blank_thickness=2, coating_material='None', coating_thickness=0,
        coating_roughness=0)

    # Arrhenius relationship values: D, Q_alpha, K, Q_K are multipled by 10_000 - Absolute values
    parameters_2 = FrictionModelParametersLiquidSolid(lambda_1=16.9633723, lambda_2=0.80269606236,
        c=0.013079927135, k_1=0.801289664383808, k_2=2.07037498, D=2.66926336189041 * 10_000,
        Q_alpha=4.98493351507642 * 10_000, k_alpha=1.00674951577508, k_3=2.81418025418379,
        k_4=1.3961773332263, K=1.0736069372 * 10_000, Q_K=7.40524741745115 * 10_000, k_5=1.857969)

    friction_model_liquid_solid = FrictionModelLiquidSolid(condition_2, parameters_2)
    friction_liquid_solid_fitter = Fitter(friction_model_liquid_solid)

    data_objects_l_s_friction = extract_workbook_data_liquid_solid_friction(workbook_name)

    chosen_experiments = {}
    for chosen_id in chosen_ids:
        chosen_experiments[chosen_id] = data_objects_l_s_friction[chosen_id]

    for dataset_object in chosen_experiments.values():
        friction_liquid_solid_fitter.add_dataset(dataset_object, verbose=True)

    ### Create parameters (Used for both manual and automatic fitting)
    friction_liquid_solid_fitter.create_params(verbose=False)

    ### Set ranges
    friction_liquid_solid_fitter.set_param_range('k_1', min=0.1, max=4.0)
    friction_liquid_solid_fitter.set_param_range('k_2', min=0.2, max=10.0)
    friction_liquid_solid_fitter.set_param_range('k_3', min=0.2, max=10.0)
    friction_liquid_solid_fitter.set_param_range('k_4', min=0.5, max=10.0)
    friction_liquid_solid_fitter.set_param_range('k_5', min=0.001, max=10.0)
    friction_liquid_solid_fitter.set_param_range('lambda_1', min=1, max=500)
    friction_liquid_solid_fitter.set_param_range('lambda_2', min=0.5, max=4.0)
    friction_liquid_solid_fitter.set_param_range('c', min=0.0001, max=10)
    friction_liquid_solid_fitter.set_param_range('k_alpha', min=0.01, max=10.0)
    friction_liquid_solid_fitter.set_param_range('D', min=1.5*10_000, max=3.0*10_000)
    friction_liquid_solid_fitter.set_param_range('K', min=0.1*10_000, max=2*10_000)
    friction_liquid_solid_fitter.set_param_range('Q_K', min=7*10_000, max=8*10_000)
    friction_liquid_solid_fitter.set_param_range('Q_alpha', min=4*10_000, max=8*10_000)

    ### Make params shared
    friction_liquid_solid_fitter.make_param_shared('D')
    friction_liquid_solid_fitter.make_param_shared('K')
    friction_liquid_solid_fitter.make_param_shared('Q_K')
    friction_liquid_solid_fitter.make_param_shared('Q_alpha')
    friction_liquid_solid_fitter.make_param_shared("k_1")
    friction_liquid_solid_fitter.make_param_shared("k_2")
    friction_liquid_solid_fitter.make_param_shared("k_3")
    friction_liquid_solid_fitter.make_param_shared("k_4")
    friction_liquid_solid_fitter.make_param_shared("k_5")
    friction_liquid_solid_fitter.make_param_shared("lambda_1")
    friction_liquid_solid_fitter.make_param_shared("lambda_2")
    friction_liquid_solid_fitter.make_param_shared("c")
    friction_liquid_solid_fitter.make_param_shared("k_alpha")

    # NOTE: Make sure that the ranges for lambda_1 and lambda_2 are reasonable because of phyiscal constraints
    # Constraint calculates lambda 1 by varying the dummy variable delta and lambda 2 while using
    # the physical constraint taking into consideration the blank roughness
    friction_liquid_solid_fitter.fit_params.add('blank_roughness', value=0.7, vary=False)
    friction_liquid_solid_fitter.fit_params.add('delta', value=10, min=5, max=20)
    friction_liquid_solid_fitter.fit_params['lambda_1_1'].expr = 'delta ** (1.0 / lambda_2_1) / blank_roughness'

    # ### Manual fitting - Will only work without usage of the numba library
    # friction_l_s_model_fitter.manual_fitting(np.linspace(0,100,100))

    ### Automatic fitting - Pick the type of optimisation method in the function
    tnc = dict(method='tnc', maxiter=10)
    dual_annealing = dict(method='dual_annealing', max_nfev=10)
    basinhopping = dict(method='basinhopping', niter=2, disp=True)
    differential_evolution = dict(method='differential_evolution', max_nfev=100000, disp=True)
    leastsq = dict(method='leastsq', maxfev=10000)

    # These new parameters will overwrite the previous new condition.
    # It will not revert to the original parameters before setting the new conditions.
    # Example: If you change lambda 1 in the first new condition and only change
    # lambda 2 in the second new condition, both lambda 1 and lambda 2 will be different
    # from the original version.

    # Start at 0, the original parameters unchanged
    # If initialise values are out of the ranges set initially, they will be brough back down
    # It isn't possible to deepcopy a jitted class at the moment
    # Always define new_values[0] as an empty dict
    new_values = {}
    new_values[0] = dict()
    # new_values[1] = dict(lambda_1=5)
    # new_values[2] = dict(lambda_1=10, c=5, k_2=2)
    # new_values[3] = dict(lambda_1=100, c=8, k_2=2, k_3=2)
    # new_values[4] = dict(lambda_1=900, c=9, k_2=2, k_3=1, k_4=2)

    results_list = []

    # Run minimise of any variant combinations of parameters
    for key, parameter_dict in new_values.items():
        print(f"Run {key}")
        for parameter, value in parameter_dict.items():
            friction_liquid_solid_fitter.fit_params[f'{parameter}_1'].set(value=value)

        result = friction_liquid_solid_fitter.automatic_fit(
            minimizer_kwargs=leastsq,
            save_results=True,
            show_plots=True
        )

        results_list.append(result)

def extract_workbook_data_liquid_solid_friction(workbook_name):
    # https://realpython.com/openpyxl-excel-spreadsheets-python/

    # One excel file with multiple sheets (e.g. results_friction_1.xlsx)
    wb = openpyxl.load_workbook(workbook_name)

    # variable_conditions: temperature, force, pressure, speed, lubricant_thickness, equiv_solid_thickness
    # x and y values: sliding_distance, coefficient_of_friction
    dataset_objects = {}

    for idx, ws in enumerate(wb.worksheets, 1):
        x_y_data = {}
        variable_conditions = {}

        zipped_variable_conditions_xl = zip(ws["D"], ws["E"])
        for cell_labels, cell_values in zipped_variable_conditions_xl:
            if cell_labels.value == None:
                break

            # Add the key value pair to the variable_conditions dictionary
            variable_conditions[cell_labels.value] = cell_values.value

        x_y_data[f"x_data"] = np.array([])
        x_y_data[f"y_data"] = np.array([])
        zipped_x_y_data_xl = zip(ws["A"], ws["B"])
        # Assume standard format of col A = SD, col B = CoF
        for col_A, col_B in zipped_x_y_data_xl:
            # Assumed that the first row is the labels and a string (e.g. COF) so it is ignored
            if isinstance(col_A.value, str):
                continue
            if col_A.value == None:
                break

            x_y_data[f"x_data"] = np.append(x_y_data[f"x_data"], col_A.value)
            x_y_data[f"y_data"] = np.append(x_y_data[f"y_data"], col_B.value)

        # Create the dataset object
        dataset_objects[idx] = (
            FrictionLiquidSolidVariableConditionsData(**variable_conditions, **x_y_data)
        )

    return dataset_objects

    # eval(f"friction_l_s_model_fitter.add_dataset(friction_solid_liquid_dataset_dict['dataset{i}'], verbose=True)")

if __name__ == "__main__":
    # fit_gaussian()
    # fit_liquid()
    fit_liquid_solid()

    # == Warnings encountered ==
    #     RuntimeWarning: overflow encountered in double_scalars
    #       H_s[i] = self.static_conditions.lubricant.avg_hardness_solid/(h_solid[i]**self.parameters.k_4)
