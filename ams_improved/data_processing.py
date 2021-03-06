import time
import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import cm
from scipy.signal import savgol_filter
from scipy.integrate import odeint
from operator import itemgetter
import sqlalchemy as db
from sqlalchemy import create_engine
from sqlalchemy import (Table, Column, Integer, Numeric, String, ForeignKey, Float,
                        PrimaryKeyConstraint, UniqueConstraint, CheckConstraint, Boolean)
from sqlalchemy import insert
from sqlalchemy import func
from sqlalchemy import update
from prettytable import PrettyTable
from pathlib import Path
import openpyxl
import shutil
import re
import sys

import database_interaction

class Experiment:
    def __init__(self, experiment_id, lubricant, pin_material, pin_roughness_Ra, blank_material, blank_roughness_Ra, blank_thickness_mm,
                 coating_material, coating_roughness_Ra, coating_thickness_mm,
                 temperature_degC, speed_mmpersecond, force_N, pressure_MPa, lubricant_thickness_micrometres):
        self.id = experiment_id

        self.conditions = ExperimentalConditions(lubricant, pin_material, pin_roughness_Ra, blank_material, blank_roughness_Ra, blank_thickness_mm,
                                                 coating_material, coating_roughness_Ra, coating_thickness_mm,
                                                 temperature_degC, speed_mmpersecond, force_N, pressure_MPa, lubricant_thickness_micrometres)

    def add_output_filename(self, filename):
        self.filename = filename

class ExperimentalConditions:
    def __init__(self, lubricant, pin_material, pin_roughness_Ra, blank_material, blank_roughness_Ra, blank_thickness_mm,
                 coating_material, coating_roughness_Ra, coating_thickness_mm,
                 temperature_degC, speed_mmpersecond, force_N, pressure_MPa, lubricant_thickness_micrometres, equiv_solid_thickness_micrometres=None):

        self.lubricant = lubricant

        self.pin_material = pin_material
        self.pin_roughness_Ra = pin_roughness_Ra

        self.blank_material = blank_material
        self.blank_roughness_Ra = blank_roughness_Ra
        self.blank_thickness_mm = blank_thickness_mm

        self.coating_material = coating_material
        self.coating_thickness_mm = coating_thickness_mm
        self.coating_roughness_Ra = coating_roughness_Ra

        self.temperature_degC = temperature_degC
        self.force_N = force_N
        self.pressure_MPa = pressure_MPa
        self.speed_mmpersecond, force_N, pressure_MPa, lubricant_thickness_micrometres = speed_mmpersecond, force_N, pressure_MPa, lubricant_thickness_micrometres
        self.lubricant_thickness_micrometres = lubricant_thickness_micrometres
        self.equiv_solid_thickness_micrometres = equiv_solid_thickness_micrometres


def sorted_nicely(l):
    """ Sort the given iterable in the way that humans expect."""
    convert = lambda text: int(text) if text.isdigit() else text
    alphanum_key = lambda key: [ convert(c) for c in re.split('([0-9]+)', key) ]
    return sorted(l, key = alphanum_key)


def extract_files_directory(load_location_raw, file_delim=','):
    """
    Gets all the files with a specific extension from the selected
    folder in the current directory. By default it will extract
    csv files.

    Parameters
    ----------
    load_location_raw: str
        The folder name containing the raw data files. It assumes that
        The file is in the same directory and the script (So this is
        just a folder name).
    file_delim: str, optional
        The separator for the file being read. By default it assumes
        that files are csv (',').

    Returns
    -------
    list
        A list containing files of the same file type. By default this
        will contain csv files in a specific folder within the current
        directory.
    """

    try:
        filenames = []
        current_dir = os.getcwd()
        path = os.path.join(current_dir, load_location_raw)
        print(f"Using specified folder path of:\n{path}")

        if file_delim==',':
            # Will use the default of .csv files
            file_delim = '.csv'
        else:
            file_delim = f'{file_delim}'

        for file in os.listdir(path):
            if file.endswith(file_delim):
                filenames.append(file)

        # Then sort them according to number
        filenames = [item for item in sorted_nicely(filenames)]

    except:
        print("Failed to extract files from the raw data folder")
        print(f"Please create a folder called \"{load_location_raw}\"")
        print("Ending program in 2 seconds")
        time.sleep(2)
        sys.exit()
    else:
        # Will return filenames if there is no exception.
        return filenames


def read_data(filenames, load_location_raw, file_delim=','):
    """
    Returns a dictionary of data frames with the filename as the key.
    Default use is with a csv file, but can be changed to tab delimited etc.

    Parameters
    ----------
    filenames: list
        A list containing the names of files that need to be read in.
    load_location_raw: str
        The folder name containing the raw data files. It assumes that
        The file is in the same directory and the script (So this is
        just a folder name).
    file_delim: str, optional
        The separator for the file being read. By default it assumes
        that files are csv (',').


    Returns
    -------
    dict
        A dictionary containing dataframes containing the raw data from
        each file passed in. The key for each dataframe will be the
        name of the file used.
    """
    dfs_raw_data = {}
    current_dir = os.getcwd()
    try:
        for csv_file in filenames:
            path = os.path.join(current_dir, load_location_raw, csv_file)
            df = pd.read_csv(path, header=None, sep=file_delim)

            # Get the csv file name (Remove the .csv)
            base = os.path.basename(csv_file)
            filename = os.path.splitext(base)[0]

            # Extract test conditions from filename
            dfs_raw_data[filename] = df
    except:
        print("Failed to read files from the raw data folder")
        print("Ending program in 2 seconds")
        time.sleep(2)
        sys.exit()

    return dfs_raw_data


def plot_scatter(df, x_axis, y_axis, test_condition=None, s=1, origin_0_0=True):
    """
    Takes in one pandas dataframe and two headers to plot a scatter.
    Titles and labels of the x axis and y axis will be determined by
    chosen dataframe headers by default.

    Call plt.show() after this is called to show the graphs.

    Parameters
    ----------
    df: pd.DataFrame
        A dataframe containing information that needs to be plotted.
    x_axis: str
        A string with the label of the column that needs to be plotted
        on the x axis.
    y_axis: str
        A string with the label of the column that needs to be plotted
        on the y axis.
    test_condition: str, optional
        A string containing the title for the graph being plotted.
    s: int, optional
        Specifies the size of each point being plotted.
    origin_0_0: bool, optional
        By default, it will make xlim and ylim at the left and bottom
        0, 0. Set this to False to disable this function.

    Returns
    -------
    None
    """
    _, ax = plt.subplots()
    ax.scatter(df[x_axis], df[y_axis], s=s)
    if test_condition == None:
        ax.set_title(f"{y_axis} vs {x_axis}")
    else:
        ax.set_title(f"{y_axis} vs {x_axis}\n{test_condition}")

    if origin_0_0==True:
        ax.set_xlim(left=0)
        ax.set_ylim(bottom=0)

    ax.set_xlabel(f"{x_axis}")
    ax.set_ylabel(f"{y_axis}")

    return None


def filter_position_data(df):
    """
    Filtering conditions for position data.

    Parameters
    ----------
    df: pd.DataFrame
        A dataframe containing position data.

    Returns
    -------
    pd.DataFrame
        A dataframe containing the only the relevant test data.
        Other data has been filtered according to set rules.
    """

    filtered_speed_range = df[(df['speed_x_(mm_s^-1)'] > 10) &
                              (df['z_position_(mm)'] < 300) &
                              (df['speed_z_(mm_s^-1)'] < 5)]

    return filtered_speed_range


def process_position_data(df, filter_data=True):
    """
    Processes the raw position data.

    Parameters
    ----------
    df: pd.DataFrame
        A dataframe containing the raw position data.
    filter_data: bool, optional
        A flag to indicate if the data should be filtered (Data collected
        outside testing will automatically be removed otherwise).

    Returns
    -------
    pd.DataFrame
        A dataframe that only contains the relevant data.
    """

    # Assign the column names of the data.
    df.drop([0], inplace=True)
    df.reset_index(drop=True, inplace=True)

    df.columns = ['time_(s)', 'x_position_(m)', 'y_position_(m)', 'z_position_(m)',
                  'x_rotation_(m)', 'y_rotation_(m)', 'z_rotation_(m)']

    df.drop(['x_rotation_(m)', 'y_rotation_(m)', 'z_rotation_(m)'], axis=1, inplace=True)
    # Convert the types of float64 (Previously objects because there was a string in the column heading)
    df = df.apply(pd.to_numeric)
    df['delta_t_(s)'] = df['time_(s)'].diff().fillna(0)

    # For position data, it is moving along the -ve x direction during the test
    df['delta_x_(m)'] = df['x_position_(m)'].diff().fillna(0)
    df['speed_x_(mm_s^-1)'] = df['delta_x_(m)'].divide(df['delta_t_(s)']).multiply(1000).abs()

    df['delta_y_(m)'] = df['y_position_(m)'].diff().fillna(0)
    df['speed_y_(mm_s^-1)'] = df['delta_y_(m)'].divide(df['delta_t_(s)']).multiply(1000).abs()

    df['delta_z_(m)'] = df['z_position_(m)'].diff().fillna(0)
    df['speed_z_(mm_s^-1)'] = df['delta_z_(m)'].divide(df['delta_t_(s)']).multiply(1000).abs()

    # Convert positions in m to mm
    df['x_position_(m)'] = df['x_position_(m)'].multiply(1000)
    df['y_position_(m)'] = df['y_position_(m)'].multiply(1000)
    df['z_position_(m)'] = df['z_position_(m)'].multiply(1000)

    # Its not done directly because for some reason it gives an error.
    df.rename(columns={'x_position_(m)': 'x_position_(mm)',
                        'y_position_(m)': 'y_position_(mm)',
                        'z_position_(m)': 'z_position_(mm)'}, inplace=True)

    # Now extract the relevant data (When the test has started)
    # If the speed is been 30 mm/s +- 5 & y and x speeds are low, it is in the correct region.

    ### Extract the required data ###
    if filter_data == True:
        filtered_speed_range = filter_position_data(df)
        # As data fluctuates a lot, extract the data from the main file where from the first time to the last time.
        # Filtering using the current method may not be reliable since sometimes the speed drops outside of the range.

        # Use -1 because otherwise its difficult to get the start captured
        start_index, end_index = filtered_speed_range.index[0], filtered_speed_range.index[-1]
        required_data = df.iloc[start_index:(end_index + 1)]
    else:
        required_data = df.copy()

    # As data fluctuates a lot, extract the data from the main file where from the first time to the last time.

    # Keep all the data now
    required_data = required_data

    return required_data


def filter_force_data(df):
    """
    Filtering conditions for force data.

    Parameters
    ----------
    df: pd.DataFrame
        A dataframe containing force data.

    Returns
    -------
    pd.DataFrame
        A dataframe containing the only the relevant test data.
        Other data has been filtered according to set rules.
    """
    processed_data = df[(df['z_force_(N)'] > 1) &
                         (df['coefficient_of_friction'] > 0) & (df['coefficient_of_friction'] < 2.5) & (df['x_force_(N)'] > -0.1) & (df['y_force_(N)'] < -0.2)].copy()

    return processed_data


def process_force_data(df, filter_data=True):
    """
    Takes in a dataframe containg raw force data, extracts the required results.
    Returns a dataframe containing: timestamp, delta t,  coefficient_of_friction,
    x, y and z forces.

    If filter_data=True then it will filter based on user set conditions.
    Otherwise it will process the whole dataframe

    Parameters
    ----------
    df: pd.DataFrame
        A dataframe containing raw force data.
    filter_data: bool, optional
        A flag to indicate if the data should be filtered (Data collected
        outside testing will automatically be removed otherwise).

    Returns
    -------
    pd.DataFrame
        A dataframe containing the processed force data.
    """

    df.columns = ['time_(s)', 'x_force_(N)', 'y_force_(N)', 'z_force_(N)']

    # Convert time column to the same scale as the position data'
    df['time_(s)'] = df['time_(s)'].apply(lambda x: x/100_000)

    df['delta_t_(s)'] = df['time_(s)'].diff().fillna(0)

    # Calculate coefficient_of_friction
    df['coefficient_of_friction'] = (df['y_force_(N)'])/(df['z_force_(N)'])

    # Convert z force to +ve values
    df['z_force_(N)'] = abs(df['z_force_(N)'])

    ### Extract the required data ###
    if filter_data == True:
        filtered_force_range = filter_force_data(df)
        start_index, end_index = filtered_force_range.index[0], filtered_force_range.index[-1]
        processed_data = df.iloc[start_index:(end_index + 1)]
    else:
        processed_data = df.copy()

    start_time = processed_data['time_(s)'].iloc[0]
    end_time = processed_data['time_(s)'].iloc[-1]
    test_time = end_time - start_time

    max_cof = processed_data['coefficient_of_friction'].max()

    return processed_data


def find_matching_csv(processed_data_keys):
    """
    Given a list of all the filenames obtained from keys in the dictionary
    storing all the extracted data by filenames, matches the corresponding force
    and position filenames and returns a list containing tuples of the
    matching filenames.

    Parameters
    ----------
    processed_data_keys: list
        A list containing filenames of all files containing the
        force and position data processed.

    Returns
    -------
    list
        A list containing tuples of matching files for force and position
        data. Each tuple with be 2 in length.
    """

    # First we need to split the data into its two parts
    # if it matches another one, with just force swapped with position or vice versa match it.
    matching_filenames = []

    # The first word is the identifier. Anything before that is extra
    for item in processed_data_keys:
        try:
            temp_split = item.split("_")
            if "force" in temp_split[0]:
                joined_temp_split = "_".join(temp_split)
                swap_force_position = joined_temp_split.replace("force", "position")
                if swap_force_position in processed_data_keys:
                    processed_data_keys.remove(swap_force_position)
                    matching_filenames.append((item, swap_force_position))
                else:
                    print(f"File is missing its matching position file:\n\t{item}")

            elif "position" in temp_split[0]:
                joined_temp_split = "_".join(temp_split)
                swap_position_force = joined_temp_split.replace("position", "force")
                if swap_position_force in processed_data_keys:
                    processed_data_keys.remove(swap_position_force)
                    matching_filenames.append((item, swap_position_force))
                else:
                    print(f"File is missing its matching force file:\n\t{item}")

            else:
                print(f"An invalid file was passed through:\n\t{item}")
        except:
            print(f"An invalid file was passed through:\n\t{item}")

    return matching_filenames


def extract_cof_results(matching_files, dfs_dictionary):
    """
    This is the last step.
    It will return only the data the user wants from selecting the points.

    Parameters
    ----------
    matching_files: list
        A list of all the matching files stored in tuples. Each tuple
        contains two strings (matching force and position file names).
    dfs_dictionary: dict
        A dictionary containing

    Returns
    -------
    dict
        A dictionary containing all the required information (in a dataframe)
        for fitting later. The key of each dataframe will be the file name.
    """
    dfs_output = {}

    for idx, match in enumerate(matching_files, 1):
        if "force" in match[0]:
            force_label = match[0]
            position_label = match[1]
        elif "position" in match[0]:
            # Normally force always comes first before position because files are stored in alphabetical order.
            # This is done just in case.
            force_label = match[1]
            position_label = match[0]
        else:
            print("Incorrect files passed in")
            return None

        # Get the test conditions to label data. This is just some string manipulation.
        key_label = match[0].replace("force_", "").replace("position_", "")
        exp_id = int(key_label.split("_")[1])

        df_force = dfs_dictionary[force_label].copy()
        df_position = dfs_dictionary[position_label].copy()

        df_force['time_elapsed_(s)'] = df_force['time_(s)'].subtract(df_force['time_(s)'].iloc[0])
        df_position['time_elapsed_(s)'] = df_position['time_(s)'].subtract(df_position['time_(s)'].iloc[0])

        df_force.drop(['time_(s)'], axis=1, inplace=True)
        df_position.drop(['time_(s)'], axis=1, inplace=True)

        # Unecessarily complex step because timestamps are not accurate
        # Match the data by using the end points
        max_time_elapsed_force = df_force['time_elapsed_(s)'].values[-1]
        max_time_elapsed_pos = df_position['time_elapsed_(s)'].values[-1]

        diff = max_time_elapsed_force - max_time_elapsed_pos

        df_force['time_elapsed_(s)'] = df_force['time_elapsed_(s)'] - diff
        df_force = df_force[df_force['time_elapsed_(s)'] > 0]

        # Making the time since test started the index column (for interpolation)
        df_force.set_index('time_elapsed_(s)', inplace=True)
        df_position.set_index('time_elapsed_(s)', inplace=True)

        # print(df_position['speed_x_(mm_s^-1)'].head())

        # Merge the two dataframes (for interpolation)
        df = pd.merge(df_force, df_position, right_index=True, left_index=True, how='outer')

        # Interpolate NaN values for columns we want (The ones we dont interpolate have NaN values)
        df['coefficient_of_friction'].interpolate(inplace=True, limit_direction='both')
        df['x_position_(mm)'].interpolate(inplace=True)
        df['x_force_(N)'].interpolate(inplace=True)
        df['speed_x_(mm_s^-1)'].interpolate(inplace=True)
        df['speed_y_(mm_s^-1)'].interpolate(inplace=True)
        df['speed_z_(mm_s^-1)'].interpolate(inplace=True)
        df['z_force_(N)'].interpolate(inplace=True)
        df['y_force_(N)'].interpolate(inplace=True)

        # Now we can reset the index since interpolation is done
        df.reset_index(inplace=True)

        # Drop duplicates caused by interpolation
        df.drop_duplicates(subset='speed_x_(mm_s^-1)', inplace=True)

        # Calculate sliding distance
        df['delta_t_(s)'] = df['time_elapsed_(s)'].diff().fillna(0)
        df['delta_x_(mm)'] = df['speed_x_(mm_s^-1)'].multiply(df['delta_t_(s)'], axis='index')
        df['sliding_distance_(mm)'] = df['delta_x_(mm)'].cumsum()

        # Always remove these columns
        df.drop(['delta_t_(s)',\
                 'delta_x_(mm)', 'delta_t_(s)_x',
                 'y_position_(mm)', 'z_position_(mm)', 'delta_t_(s)_y',
                 'delta_x_(m)', 'delta_y_(m)', 'speed_y_(mm_s^-1)',
                 'delta_z_(m)', 'speed_z_(mm_s^-1)'], axis=1, inplace=True)

        # Drop any duplicates in time column potentially due to interpolation
        df.drop_duplicates(subset='time_elapsed_(s)', inplace=True)

        # Remove the first row that has empty cells (and any other ones)
        df.dropna(inplace=True)

        min_time = 0.06
        df = df[df['time_elapsed_(s)'] > min_time]
        df['coefficient_of_friction'] = savgol_filter(df['coefficient_of_friction'].values, 51, 2)

        dfs_output[exp_id] = df

    return dfs_output


def check_folder_exists(folder_path):
    """
    Checks if the folder exists. If the folder doesn't exist, it is
    created.

    Parameters
    ----------
    folder_path: str
        An string with the name of the folder that is being checked
        to see if it exists.

    Returns
    -------
    None
    """
    if os.path.exists(folder_path):
        pass
    else:
        print(f"- Folder {folder_path} doesn't exist! Creating a new folder.")
        try:
            os.mkdir(folder_path)
        except OSError:
            print (f"Creation of the directory {folder_path} failed")
        else:
            print (f"Created folder: {folder_path} ")

    return None


def process_force_position(dfs_raw_data):
    """
    Parameters
    ----------
    dfs_raw_data: dict
        A dictionary containing pandas DataFrames with keys as the file
        names. Testing raw data (position + force) is stored in each DataFrame.

    Returns
    -------
    dict
        A dictionary of the same format as the input dictionary except now
        the raw data (force + position) in each dataframe has been
        processed and filtered as required.
    """
    # Initialise the dictionary containing the full processed data (Before user input)
    dfs_position_force_filtered = {}
    count = 0

    for file_name, df in dfs_raw_data.items():

        if "force" in file_name:
            count = count + 1
            print(f"Files processed: {count}", end='\r')


            # Filtered Data
            force_data = process_force_data(df, filter_data=True)
            dfs_position_force_filtered[file_name] = force_data

        elif "position" in file_name:
            count = count + 1
            print(f"Files processed: {count}", end='\r')

            # Filtered Data
            position_data = process_position_data(df, filter_data=True)
            dfs_position_force_filtered[file_name] = position_data

        else:

            print(f"Incorrect File Passed Through.\n")


    print(f"Files processed: {count}")

    return dfs_position_force_filtered


def main():
    """ Main Function for running the script. """
    t0 = time.time()

    ### USER INPUT ###
    # Indicate the relative location of raw data and bins folders
    raw_data_folder = "friction_raw_data"
    bin_folder = "friction_raw_data_bin"

    # Assumption that the experiment number of the raw data files are
    #   accurate and match that of the excel worksheet.
    # Test conditions are obtained from the excel worksheet.
    filename_exp_info = "_friction_model_1.xlsx"

    exp_info_filepath = Path(filename_exp_info).resolve()
    ### USER INPUT ###

    startup_text = "Processing Files for Interactive Friction Model."
    print(startup_text)

    # # Initialisation for testing
    # source1 = r"C:\Users\ismael_zainal\Desktop\repos\fyp-2020\ams_improved\friction_raw_data\hidden\force_data_1.csv"
    # source2 = r"C:\Users\ismael_zainal\Desktop\repos\fyp-2020\ams_improved\friction_raw_data\hidden\position_data_1.csv"

    # destination1 = r"C:\Users\ismael_zainal\Desktop\repos\fyp-2020\ams_improved\friction_raw_data\force_data_1.csv"
    # destination2 = r"C:\Users\ismael_zainal\Desktop\repos\fyp-2020\ams_improved\friction_raw_data\position_data_1.csv"

    # shutil.copy(source1, destination1)
    # shutil.copy(source2, destination2)

    ################## Step 1 - Check if the required folders exist ##################

    print("="*len(startup_text))
    print("Checking if the folder for raw data exists...")
    check_folder_exists(raw_data_folder)

    print("Checking if the bin folder exists...")
    check_folder_exists(bin_folder)
    print("="*len(startup_text))

    ################## Step 2 - Extract the csv files to be processed ##################

    csv_files = extract_files_directory(load_location_raw=raw_data_folder)
    dfs_raw_data = read_data(csv_files, load_location_raw=raw_data_folder)
    print("="*len(startup_text))

    ################## Step 3 - Process all the force and position data ##################

    # print("Processing Force and Position Files...")

    dfs_position_force_filtered = process_force_position(dfs_raw_data)

    t1 = time.time()
    time_processing_data = round((t1-t0),4)
    print(f"Time to process force and position data is: {time_processing_data}s")

    ################## Step 4 - Match all the position and force files ##################

    list_keys = list(dfs_position_force_filtered.keys())
    matching_files = find_matching_csv(list_keys)

    ################## Step 5 - Obtain the coefficient_of_friction graphs ################

    # print("Calculating Coefficient of Friction vs Sliding Distance...")

    dfs_result = extract_cof_results(matching_files, dfs_position_force_filtered)

    ################## Step 6 - Plot the extracted data #################################

    t2 = time.time()
    print(f"Time to calculate coefficient of friction: {round((t2-t1), 4)}s\n")

    print(f"Output file plot:")
    for experiment_id in dfs_result.keys():
        print(f"Experimental Result {experiment_id}")

        ### PLOTTING
        plot_scatter(dfs_result[experiment_id], 'sliding_distance_(mm)', 'coefficient_of_friction', test_condition=f'Experiment: {experiment_id}')
        # The showing of the graphs have been disabled for now
        # plt.show()

    ################## Step 7 - Create results file in results folder & move processed data to bin directory ##################################
    # Add experiment to database
    # Get the new experiment number
    # Rename the output number
    subdir = 'results_friction'
    print("")
    print("Checking if results folder exists...")
    check_folder_exists(subdir)
    # create the files

    for file1, file2 in matching_files:
        # Get the experiment id
        local_experiment_id = int(file1.split("_")[2])

        wb = openpyxl.load_workbook(exp_info_filepath)
        ws = wb.worksheets[0]

        # Assert all column names are what we expect
        assert ws['A1'].value == "experiment_id"
        assert ws['B1'].value == "lubricant_name"
        assert ws['C1'].value == "pin_material"
        assert ws['D1'].value == "pin_roughness_Ra"
        assert ws['E1'].value == "blank_material"
        assert ws['F1'].value == "blank_roughness_Ra"
        assert ws['G1'].value == "blank_thickness_mm"
        assert ws['H1'].value == "coating_material"
        assert ws['I1'].value == "coating_roughness_Ra"
        assert ws['J1'].value == "coating_thickness_mm"
        assert ws['K1'].value == "temperature_degC"
        assert ws['L1'].value == "speed_mmpersecond"
        assert ws['M1'].value == "force_N"
        assert ws['N1'].value == "pressure_MPa"
        assert ws['O1'].value == "lubricant_thickness_micrometres"
        assert ws['P1'].value == "processed?"

        local_experiment_row = 0
        for idx, row in enumerate(ws.iter_rows(), 1):
            # experiment_id will always be the first row
            # processed? Y/N will always be the last row
            if ws[f'A{idx}'].value == local_experiment_id:
                if ws[f'P{idx}'].value == "y":
                    raise Exception("Experiment has already been processed according to excel worksheet, please check data filenames or correct the sheet")
                else:
                    local_experiment_row = idx
                    experiment = Experiment(row[0].value, row[1].value, row[2].value, row[3].value,
                               row[4].value, row[5].value, row[6].value, row[7].value,
                               row[8].value, row[9].value, row[10].value, row[11].value,
                               row[12].value, row[13].value, row[14].value)

                    try:
                        ws[f'P{local_experiment_row}'].value = "y"
                        wb.save(exp_info_filepath)
                    except PermissionError as msg:
                        raise Exception(f"PermissionsError, please ensure that the file {exp_info_filepath} is not open")

                    database_interaction.add_experiment_database(experiment)

                    # Get the output filename (according to database)
                    db_experiment_id, proc_data_filename = database_interaction.get_last_experiment_info()

                    break

        output_filepath = Path(subdir).joinpath(proc_data_filename)

        dfs_result[local_experiment_id].to_csv(output_filepath, sep=',', index=False)

        # Move the processed files to the raw data bin & rename according to db numbering
        raw_data_path1 = f"{raw_data_folder}\\{file1}.csv"
        raw_data_path2 = f"{raw_data_folder}\\{file2}.csv"

        # Assert that the filename is as expected so we can use an _ to separate the number
        # This is the filename without the csv
        filename_pattern = '(force|position)_data_([0-9]+)'
        filename_matcher = re.compile(filename_pattern)

        matcher1 = filename_matcher.match(file1)
        matcher2 = filename_matcher.match(file2)

        assert matcher1 is not None
        assert matcher2 is not None

        new_filename1 = f"{matcher1.group(1)}_data_{db_experiment_id}"
        new_filename2 = f"{matcher2.group(1)}_data_{db_experiment_id}"

        bin_path1 = f"{bin_folder}\\{new_filename1}.csv"
        bin_path2 = f"{bin_folder}\\{new_filename2}.csv"

        shutil.move(raw_data_path1, bin_path1)
        shutil.move(raw_data_path2, bin_path2)

if __name__ == "__main__":
    # Setup Code Runner to allow user input:
    # https://stackoverflow.com/questions/50689210/how-to-setup-code-runner-in-visual-studio-code-for-python

    # https://www.youtube.com/watch?v=k27MJJLJNT4 Maybe try holoview. It allows you to easily overlay a line.

    # For plotting:
    # https://matplotlib.org/gallery/index.html#our-favorite-recipes

    # try:
    #     # remove existing database
    #     os.remove('friction_model.db')
    # except:
    #     raise Exception("failed to remove db")

    # create the database if it doesn't exist
    db_filename = "friction_model.db"
    if not Path(db_filename).exists():
        print("No database detected, initialising a new one")
        database_interaction.create_database()
    else:
        print("Existing database detected")

    main()
